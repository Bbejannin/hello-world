import os as os
from xlwt import Workbook
import xlrd
from collections import defaultdict
import xlwt
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side

book = openpyxl.Workbook()
wb = xlrd.open_workbook('Déplacements_maintenance.xlsx')
thin_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
blueFill = PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid')
greenFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
greyFill = PatternFill(start_color='FFC0C0C0', end_color='FFC0C0C0', fill_type='solid')

sheet1 = book.active
sheet1.cell(row=1, column=1).value = "Entreprise"
sheet1.column_dimensions['A'].width = 31.77
sheet1.cell(row = 1, column = 1).font = Font(bold = True) 
sheet1['A1'].alignment = Alignment(horizontal='center')
sheet1.cell(row=1, column=1).border = thin_border
sheet1['A1'].fill = greyFill
sheet1.cell(row=1, column=2).value = "Coût déplacement"
sheet1.column_dimensions['B'].width = 17.5
sheet1.cell(row = 1, column = 2).font = Font(bold = True) 
sheet1['B1'].alignment = Alignment(horizontal='center')
sheet1.cell(row=1, column=2).border = thin_border
sheet1['B1'].fill = greyFill
sheet1.cell(row=1, column=3).value = "Nbre déplacements"
sheet1.column_dimensions['C'].width = 18.6
sheet1.cell(row = 1, column = 3).font = Font(bold = True) 
sheet1['C1'].alignment = Alignment(horizontal='center')
sheet1.cell(row=1, column=3).border = thin_border
sheet1['C1'].fill = greyFill
sh = wb.sheet_by_index(0)

cost_max = 400 #Cost max (€)
nb_min = 2 #Nombre minimum de dépannages.

companies_cost = defaultdict(int)
companies_count = defaultdict(int)
for i in range(sh.nrows-1):
    name_company = sh.row_values(i+1)[7]
    companies_cost[name_company] += float(sh.row_values(i+1)[2])
    companies_count[name_company] += 1
nb_l = 2
for name, _ in companies_cost.items():
    companies_cost[name] /= companies_count[name]
    #print(str(name)+' - '+str(round(companies_cost[name],2))+' € ('+str(companies_count[name])+')')
    cell_A = 'A'+str(nb_l)
    cell_B = 'B'+str(nb_l)
    cell_C = 'C'+str(nb_l)
    sheet1[cell_A].alignment = Alignment(horizontal='center')
    sheet1.cell(row=nb_l, column=1).border = thin_border
    sheet1[cell_B].alignment = Alignment(horizontal='center')
    sheet1.cell(row=nb_l, column=2).border = thin_border
    sheet1[cell_C].alignment = Alignment(horizontal='center')
    sheet1.cell(row=nb_l, column=3).border = thin_border
    if companies_cost[name] < cost_max and companies_count[name] > nb_min:
        sheet1.cell(row=nb_l, column=1).value = name
        sheet1[cell_A].fill = yellowFill
    else:
        sheet1.cell(row=nb_l, column=1).value = name  
    if companies_cost[name] < cost_max:
        sheet1.cell(row=nb_l, column=2).value = round(companies_cost[name],2)#, st)
        sheet1[cell_B].fill = blueFill
    else:
        sheet1.cell(row=nb_l, column=2).value = round(companies_cost[name],2)  
    if companies_count[name] > nb_min:
        sheet1.cell(row=nb_l, column=3).value = companies_count[name]#, st)
        sheet1[cell_C].fill = greenFill
    else:
        sheet1.cell(row=nb_l, column=3).value = companies_count[name]
    nb_l += 1
book.save("companies_cost.xlsx")
os.startfile("companies_cost.xlsx")